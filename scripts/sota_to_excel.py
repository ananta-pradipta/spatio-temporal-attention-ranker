"""Convert SOTA landscape markdown to Excel with multiple sheets."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_md_table(lines):
    """Parse markdown table lines into list of dicts."""
    rows = []
    headers = None
    for line in lines:
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if not cells:
            continue
        # Skip separator rows
        if all(set(c) <= set("- :") for c in cells):
            continue
        if headers is None:
            headers = cells
        else:
            rows.append(dict(zip(headers, cells)))
    return headers, rows


def add_sheet(wb, name, headers, rows):
    """Add a formatted sheet to workbook."""
    ws = wb.create_sheet(title=name[:31])  # Excel 31-char limit

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Write data
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            # Clean markdown bold
            val = val.replace("**", "")
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-width (approximate)
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def main():
    md_path = "/home/apradipta/phd-research/literature/sota-landscape.md"
    out_path = "/home/apradipta/phd-research/literature/sota-landscape.xlsx"

    with open(md_path) as f:
        content = f.read()

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define sections to extract (heading -> sheet name)
    sections = [
        ("## Overall Priority Rank", "Overall Rank"),
        ("### Foundational Methods", "T-GNN Foundational"),
        ("### Financial Applications", "T-GNN Financial"),
        ("## Axis 2: Financial NLP", "Financial NLP"),
        ("## Axis 3: Graph Construction", "Graph Construction"),
        ("## Axis 4: Agentic AI", "Agentic AI"),
        ("## Axis 5: RL plus GNN", "RL + GNN Portfolio"),
        ("## Cross-Axis Gap Analysis", "Gap Analysis"),
    ]

    lines = content.split("\n")

    for heading, sheet_name in sections:
        # Find the section
        start = None
        for i, line in enumerate(lines):
            if heading in line:
                start = i
                break
        if start is None:
            continue

        # Collect lines until next heading of same or higher level
        heading_level = len(heading) - len(heading.lstrip("#"))
        table_lines = []
        for i in range(start + 1, len(lines)):
            line = lines[i]
            # Check if we hit a new section of same or higher level
            if line.startswith("#"):
                lvl = len(line) - len(line.lstrip("#"))
                if lvl <= heading_level and line.strip("# \n"):
                    break
            if line.strip().startswith("|"):
                table_lines.append(line)

        if table_lines:
            headers, rows = parse_md_table(table_lines)
            if headers and rows:
                add_sheet(wb, sheet_name, headers, rows)

    if wb.sheetnames:
        wb.save(out_path)
        print(f"Saved {len(wb.sheetnames)} sheets to {out_path}")
    else:
        print("No tables found")


if __name__ == "__main__":
    main()
